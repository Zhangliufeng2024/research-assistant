"""Deterministic input fingerprints and lightweight tabular schema snapshots."""

from __future__ import annotations

import csv
import hashlib
import importlib.metadata
import json
import platform
import re
import sys
from pathlib import Path
from typing import Any

SUPPORTED_SCHEMA_SUFFIXES = {".csv", ".tsv", ".json", ".parquet", ".xlsx", ".xls"}


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _value_type(value: Any) -> str:
    if value is None or value == "":
        return "null"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, int) and not isinstance(value, bool):
        return "int"
    if isinstance(value, float):
        return "float"
    if isinstance(value, (dict, list)):
        return "object" if isinstance(value, dict) else "array"
    text = str(value).strip()
    try:
        int(text)
        return "int"
    except ValueError:
        try:
            float(text)
            return "float"
        except ValueError:
            return "string"


def _merge_types(values: list[Any]) -> str:
    types = {_value_type(value) for value in values if _value_type(value) != "null"}
    if not types:
        return "null"
    if types <= {"int", "float"}:
        return "float" if "float" in types else "int"
    return next(iter(types)) if len(types) == 1 else "mixed"


def schema_for_file(path: str | Path) -> dict[str, Any]:
    target = Path(path)
    suffix = target.suffix.lower()
    result: dict[str, Any] = {
        "path": str(target), "suffix": suffix, "sha256": file_sha256(target),
        "size": target.stat().st_size,
    }
    if suffix in {".csv", ".tsv"}:
        with target.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
            reader = csv.reader(handle, delimiter="\t" if suffix == ".tsv" else ",")
            header = next(reader, [])
            samples = [row for _, row in zip(range(1000), reader, strict=False)]
        result.update({"kind": "table", "columns": header, "types": [_merge_types([row[i] for row in samples if i < len(row)]) for i in range(len(header))], "sample_rows": len(samples)})
        return result
    if suffix == ".json":
        decoded = json.loads(target.read_text(encoding="utf-8"))
        records = decoded if isinstance(decoded, list) else [decoded]
        records = [row for row in records if isinstance(row, dict)]
        columns = sorted({key for row in records for key in row})
        result.update({"kind": "json", "shape": "array" if isinstance(decoded, list) else "object", "columns": columns, "types": {key: _merge_types([row.get(key) for row in records]) for key in columns}, "records": len(records)})
        return result
    if suffix in {".xlsx", ".xls"}:
        import openpyxl
        workbook = openpyxl.load_workbook(target, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True, max_row=1001))
        header = [str(value) if value is not None else "" for value in (rows[0] if rows else [])]
        samples = [list(row) for row in rows[1:]]
        result.update({"kind": "table", "sheet": sheet.title, "columns": header, "types": [_merge_types([row[i] for row in samples if i < len(row)]) for i in range(len(header))], "sample_rows": len(samples)})
        workbook.close()
        return result
    if suffix == ".parquet":
        import pandas as pd
        frame = pd.read_parquet(target)
        result.update({"kind": "table", "columns": [str(column) for column in frame.columns], "types": {str(column): str(dtype) for column, dtype in frame.dtypes.items()}, "rows": int(len(frame))})
        return result
    result["kind"] = "file"
    return result


def snapshot_input_files(root: str | Path, paths: list[str | Path]) -> dict[str, Any]:
    workspace = Path(root).resolve()
    files: list[dict[str, Any]] = []
    schemas: dict[str, Any] = {}
    for raw in paths:
        target = Path(raw)
        if not target.is_absolute():
            target = workspace / target
        target = target.resolve()
        try:
            rel = target.relative_to(workspace).as_posix()
        except ValueError:
            continue
        if not target.is_file():
            continue
        record = {"path": rel, "sha256": file_sha256(target), "size": target.stat().st_size}
        files.append(record)
        if target.suffix.lower() in SUPPORTED_SCHEMA_SUFFIXES:
            try:
                schemas[rel] = schema_for_file(target)
            except Exception as exc:  # malformed optional data should be visible, not fatal
                schemas[rel] = {"path": rel, "sha256": record["sha256"], "error": str(exc)[:500]}
    return {"files": files, "schemas": schemas}


def schema_changes(previous: dict[str, Any], current: dict[str, Any]) -> list[dict[str, Any]]:
    changes: list[dict[str, Any]] = []
    paths = sorted(set(previous) | set(current))
    for path in paths:
        if previous.get(path) != current.get(path):
            changes.append({"path": path, "before": previous.get(path), "after": current.get(path)})
    return changes


def environment_lock(root: str | Path | None = None) -> dict[str, Any]:
    """Build a compact, deterministic lock for reproducible local runs.

    This is intentionally based on project-declared distributions rather than
    ``pip freeze``: it is portable across virtualenvs, avoids recording secrets
    or editable checkout paths, and still catches dependency drift. A caller
    may persist the returned lock beside an analysis manifest.
    """
    workspace = Path(root).resolve() if root else Path.cwd().resolve()
    pyproject = workspace / "pyproject.toml"
    source_hash = file_sha256(pyproject) if pyproject.is_file() else ""
    declared: list[str] = []
    if pyproject.is_file():
        try:
            text = pyproject.read_text(encoding="utf-8")
            try:
                import tomllib
                document = tomllib.loads(text)
                project = document.get("project", {}) if isinstance(document, dict) else {}
                raw_dependencies = list(project.get("dependencies") or [])
                optional = project.get("optional-dependencies") or {}
                for values in optional.values() if isinstance(optional, dict) else []:
                    raw_dependencies.extend(values or [])
            except ModuleNotFoundError:  # Python 3.10 without tomli installed
                raw_dependencies = re.findall(
                    r'^\s*"([A-Za-z][A-Za-z0-9_.-]*(?:\[[^]]+\])?(?:[<>=!~].*)?)"',
                    text, re.MULTILINE,
                )
            for raw in raw_dependencies:
                name = str(raw).split("[", 1)[0]
                for marker in ("<", ">", "=", "!", "~", ";"):
                    name = name.split(marker, 1)[0]
                normalized = name.strip().lower().replace("_", "-")
                if normalized and normalized not in declared:
                    declared.append(normalized)
        except (OSError, ValueError, TypeError, AttributeError):
            # A malformed or legacy project manifest should not prevent a run;
            # the pyproject hash still records that the lock source changed.
            declared = []
    packages: dict[str, str] = {}
    for name in sorted(declared):
        try:
            packages[name] = importlib.metadata.version(name)
        except importlib.metadata.PackageNotFoundError:
            packages[name] = "<missing>"
    lock_payload = {"pyproject_sha256": source_hash, "packages": packages}
    lock_hash = hashlib.sha256(
        json.dumps(lock_payload, sort_keys=True, separators=(",", ":")).encode("utf-8")
    ).hexdigest()
    return {"lock_hash": lock_hash, **lock_payload}


def runtime_environment(root: str | Path | None = None) -> dict[str, Any]:
    lock = environment_lock(root)
    return {
        "python": platform.python_version(),
        "platform": platform.platform(),
        "executable": sys.executable,
        "dependency_lock_hash": lock["lock_hash"],
        "dependency_lock": lock,
    }
